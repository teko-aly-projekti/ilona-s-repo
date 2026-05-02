from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from helpers import (
    classify_category,
    find_month_total_from_worksheet,
    is_lab_row,
    parse_month_from_sheet,
    read_clean_rows,
)


EXCEL_FILE = "tamperetalo2023.xlsx"
OUTPUT_DIR = "output"


def main():
    out = Path(OUTPUT_DIR)
    out.mkdir(exist_ok=True)

    df = read_clean_rows(EXCEL_FILE)
    df["Kategoria"] = df.apply(lambda r: classify_category(r["Seloste"], r["Koodi"]), axis=1)
    df["Onko_labra"] = df.apply(lambda r: is_lab_row(r["Seloste"], r["Koodi"]), axis=1)

    wb = load_workbook(EXCEL_FILE, data_only=True)
    month_rows = []
    for sheet_name in wb.sheetnames:
        total = find_month_total_from_worksheet(wb[sheet_name], sheet_name)
        month_rows.append({"Kuukausi": parse_month_from_sheet(sheet_name), "Summa €": total})

    monthly = pd.DataFrame(month_rows).sort_values("Kuukausi").reset_index(drop=True)

    total_year = monthly["Summa €"].sum()
    peak_month_row = monthly.loc[monthly["Summa €"].idxmax()]
    low_month_row = monthly.loc[monthly["Summa €"].idxmin()]

    category_totals = (
        df.groupby("Kategoria")["Summa €"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )

    top_category = category_totals.iloc[0]

    top_item = (
        df.groupby(["Koodi", "Seloste"])["Summa €"]
        .sum()
        .reset_index()
        .sort_values("Summa €", ascending=False)
        .iloc[0]
    )

    lab_total = df[df["Onko_labra"]]["Summa €"].sum()
    lab_share = (lab_total / total_year * 100) if total_year else 0

    monthly_by_cat = (
        df.groupby(["Kuukausi", "Kategoria"])["Summa €"]
        .sum()
        .reset_index()
        .pivot(index="Kuukausi", columns="Kategoria", values="Summa €")
        .fillna(0)
        .sort_index()
    )

    cat_growth = (monthly_by_cat.iloc[-1] - monthly_by_cat.iloc[0]).sort_values(ascending=False)
    fastest_category = cat_growth.index[0]
    fastest_growth = cat_growth.iloc[0]

    insights = [
        f"Vuoden kokonaiskustannus oli {total_year:.2f} euroa.",
        f"Korkein kustannuskuukausi oli {peak_month_row['Kuukausi']} ({peak_month_row['Summa €']:.2f} euroa).",
        f"Matalin kustannuskuukausi oli {low_month_row['Kuukausi']} ({low_month_row['Summa €']:.2f} euroa).",
        f"Suurin kulukategoria oli {top_category['Kategoria']} ({top_category['Summa €']:.2f} euroa).",
        f"Suurin yksittäinen kustannusrivi oli {top_item['Koodi']} | {top_item['Seloste']} ({top_item['Summa €']:.2f} euroa).",
        f"Laboratoriokulujen osuus koko vuoden kustannuksista oli {lab_share:.1f} %.",
        f"Nopeimmin kasvanut kategoria alkuvuoden ja loppuvuoden välillä oli {fastest_category} ({fastest_growth:.2f} euroa).",
    ]

    recommendations = [
        "Seuratkaa erityisesti loppuvuoden kustannuspiikkejä, koska syksyllä kokonaiskulut nousivat selvästi.",
        "Kohdistakaa tarkempi selvitys suurimpaan kulukategoriaan ja suurimpiin yksittäisiin kustannusriveihin.",
        "Jos etäpalveluiden tai mielenterveyspalveluiden kasvu jatkuu, arvioikaa kuormituksen ja varhaisen tuen toimintamalleja.",
        "Laboratoriokuluista kannattaa tarkistaa, mitkä tutkimukset toistuvat usein ja mitkä liittyvät mahdollisiin pidempikestoisiin työkykyteemoihin.",
    ]

    lines = [
        "# Työterveysdatan insight-yhteenveto",
        "",
        "## Keskeiset havainnot",
    ]
    lines.extend([f"- {item}" for item in insights])
    lines.extend(["", "## Suositeltavat jatkotoimet"])
    lines.extend([f"- {item}" for item in recommendations])

    report_path = out / "08_insight_yhteenveto.md"
    report_path.write_text("\n".join(lines), encoding="utf-8")

    summary_table = pd.DataFrame(
        {
            "Mittari": [
                "Vuoden kokonaiskustannus",
                "Korkein kustannuskuukausi",
                "Matalin kustannuskuukausi",
                "Suurin kulukategoria",
                "Suurin yksittäinen kustannusrivi",
                "Labrojen osuus %",
                "Nopeimmin kasvanut kategoria",
            ],
            "Arvo": [
                f"{total_year:.2f} €",
                f"{peak_month_row['Kuukausi']} ({peak_month_row['Summa €']:.2f} €)",
                f"{low_month_row['Kuukausi']} ({low_month_row['Summa €']:.2f} €)",
                f"{top_category['Kategoria']} ({top_category['Summa €']:.2f} €)",
                f"{top_item['Koodi']} | {top_item['Seloste']} ({top_item['Summa €']:.2f} €)",
                f"{lab_share:.1f}",
                f"{fastest_category} ({fastest_growth:.2f} €)",
            ],
        }
    )
    summary_table.to_csv(out / "08_insight_yhteenveto.csv", index=False)

    print(summary_table)
    print(f"Valmis: {report_path}")


if __name__ == "__main__":
    main()
