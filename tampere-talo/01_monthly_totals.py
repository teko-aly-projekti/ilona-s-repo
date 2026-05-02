from __future__ import annotations

from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook

from helpers import find_month_total_from_worksheet, parse_month_from_sheet


EXCEL_FILE = "tamperetalo2023.xlsx"
OUTPUT_DIR = "output"


def main():
    path = Path(EXCEL_FILE)
    out = Path(OUTPUT_DIR)
    out.mkdir(exist_ok=True)

    wb = load_workbook(path, data_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total = find_month_total_from_worksheet(ws, sheet_name)
        rows.append({"Kuukausi": parse_month_from_sheet(sheet_name), "Summa €": total})

    monthly = pd.DataFrame(rows).sort_values("Kuukausi").reset_index(drop=True)
    print(monthly)

    sns.set_theme(style="whitegrid", context="talk")
    fig, ax = plt.subplots(figsize=(14, 7))

    sns.barplot(data=monthly, x="Kuukausi", y="Summa €", color="#2E6F95", ax=ax)

    ax.set_title("Työterveyden kuukausikustannukset 2023", fontsize=20, pad=20)
    ax.set_xlabel("Kuukausi")
    ax.set_ylabel("Euroa")
    ax.tick_params(axis="x", rotation=30)

    for i, value in enumerate(monthly["Summa €"]):
        ax.text(i, value + 50, f"{value:.2f} €", ha="center", va="bottom", fontsize=11)

    plt.tight_layout()
    fig.savefig(out / "01_kuukausikustannukset.png", dpi=220, bbox_inches="tight")
    monthly.to_csv(out / "01_kuukausikustannukset.csv", index=False)

    print(f"Valmis: {out / '01_kuukausikustannukset.png'}")


if __name__ == "__main__":
    main()
